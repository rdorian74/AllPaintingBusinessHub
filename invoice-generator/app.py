from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
import os
import json
import re
from datetime import datetime
from weasyprint import HTML, CSS
import tempfile
import pdfplumber

app = Flask(__name__, static_folder='static')
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir() + '/invoice_uploads'
app.config['OUTPUT_FOLDER'] = tempfile.gettempdir() + '/invoice_outputs'
app.config['DATA_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
os.makedirs(app.config['DATA_FOLDER'], exist_ok=True)

# Check if logo exists
LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'logo.png')
HAS_LOGO = os.path.exists(LOGO_PATH)
# Convert to proper file URL for Windows
if HAS_LOGO:
    LOGO_URL = LOGO_PATH.replace('\\', '/')
    if not LOGO_URL.startswith('/'):
        LOGO_URL = '/' + LOGO_URL
else:
    LOGO_URL = None

# Invoice counter file
COUNTER_FILE = os.path.join(app.config['DATA_FOLDER'], 'invoice_counter.json')


def get_current_year_prefix():
    """Get the year prefix for invoice numbers (e.g., '25C' for 2025)"""
    year = datetime.now().year
    return f"{str(year)[-2:]}C"


def load_invoice_counter():
    """Load the invoice counter from file"""
    if os.path.exists(COUNTER_FILE):
        with open(COUNTER_FILE, 'r') as f:
            data = json.load(f)
            current_prefix = get_current_year_prefix()
            if data.get('year_prefix') != current_prefix:
                return {'year_prefix': current_prefix, 'last_number': 0, 'sequences': {}}
            return data
    return {'year_prefix': get_current_year_prefix(), 'last_number': 0, 'sequences': {}}


def save_invoice_counter(data):
    """Save the invoice counter to file"""
    with open(COUNTER_FILE, 'w') as f:
        json.dump(data, f, indent=2)


def get_next_invoice_number(estimate_prefix=None):
    """
    Generate the next invoice number.
    Format: XXY-NNN where:
    - XXY is the estimate prefix (e.g., 25R, 25C, 26C)
    - NNN is the sequential invoice number for that prefix
    """
    counter = load_invoice_counter()
    
    # Use the estimate prefix if provided, otherwise use current year
    if estimate_prefix:
        prefix = estimate_prefix
    else:
        prefix = counter['year_prefix']
    
    sequences = counter.get('sequences', {})
    
    # Get next sequence for this prefix
    current_seq = sequences.get(prefix, 0)
    next_seq = current_seq + 1
    
    # Update counter
    sequences[prefix] = next_seq
    counter['sequences'] = sequences
    save_invoice_counter(counter)
    
    return f"{prefix}-{next_seq:03d}"


def safe_float(value, default=0.0):
    """Safely convert a value to float"""
    if value is None or value == '':
        return default
    try:
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '').strip()
        return float(value)
    except (ValueError, TypeError):
        return default


def extract_data_from_estimate_pdf(pdf_path):
    """Extract client info, project info, and scope items from estimate PDF page 1
    Handles both NEW format (from estimate generator) and LEGACY format (old Word docs)
    """
    
    extracted = {
        'estimate_number': '',
        'estimate_prefix': '',
        'base_number': 0,
        'date': datetime.now().strftime('%B %d, %Y'),
        'client_company': '',
        'client_contact': '',
        'client_address': '',
        'client_phone': '',
        'client_email': '',
        'project_name': '',
        'project_address': '',
        'line_items': [],
        'subtotal': 0.0
    }
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if len(pdf.pages) > 0:
                page = pdf.pages[0]
                text = page.extract_text() or ''
                
                print(f"\n{'='*60}")
                print("EXTRACTING DATA FROM ESTIMATE PDF")
                print(f"{'='*60}")
                print(f"Page 1 text length: {len(text)} characters")
                
                # Detect format type
                is_legacy_format = 'OUR ESTIMATE #' in text or 'Project Scope #' in text
                
                if is_legacy_format:
                    print("📋 Detected LEGACY format estimate")
                    extracted = extract_legacy_format(text, extracted)
                else:
                    print("📋 Detected NEW format estimate")
                    extracted = extract_new_format(text, page, extracted)
                
                print(f"{'='*60}\n")
                
    except Exception as e:
        print(f"❌ Error extracting PDF: {e}")
        import traceback
        traceback.print_exc()
    
    return extracted


def extract_legacy_format(text, extracted):
    """Extract data from legacy/old format estimates (Word doc style)"""
    
    lines = text.split('\n')
    
    # Extract estimate number (e.g., "OUR ESTIMATE # 25R– 281 REVISED")
    est_match = re.search(r'OUR ESTIMATE\s*#?\s*(\d{2}[A-Z]?\s*[-–]\s*\d+)', text, re.IGNORECASE)
    if est_match:
        extracted['estimate_number'] = est_match.group(1).replace('–', '-').replace(' ', '')
        # Extract prefix (e.g., "25R") and base number
        prefix_match = re.search(r'(\d{2}[A-Z]?)-(\d+)', extracted['estimate_number'])
        if prefix_match:
            extracted['estimate_prefix'] = prefix_match.group(1)
            extracted['base_number'] = int(prefix_match.group(2))
        print(f"✅ Estimate number: {extracted['estimate_number']}")
        print(f"✅ Estimate prefix: {extracted.get('estimate_prefix', 'N/A')}")
    
    # Extract client name (usually first line after header, or line with name before date)
    for i, line in enumerate(lines[:10]):
        if any(skip in line.upper() for skip in ['SITE COPY', 'ALL PAINTING', 'SCHOOLHOUSE', 'WWW.', 'FULL-SERVICE']):
            continue
        name_match = re.match(r'^([A-Za-z][A-Za-z\s]+?)(?:\s{2,}|\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))', line)
        if name_match:
            extracted['client_contact'] = name_match.group(1).strip()
            extracted['client_company'] = extracted['client_contact']
            print(f"✅ Client name: {extracted['client_contact']}")
            break
        elif re.match(r'^[A-Z][a-z]+\s+[A-Z][a-z]+$', line.strip()):
            extracted['client_contact'] = line.strip()
            extracted['client_company'] = extracted['client_contact']
            print(f"✅ Client name: {extracted['client_contact']}")
            break
    
    # Extract address (line with street number)
    for i, line in enumerate(lines[:15]):
        # Clean the line - remove any date info that might be on same line
        clean_line = re.sub(r'\s+(January|February|March|April|May|June|July|August|September|October|November|December).*$', '', line.strip(), flags=re.IGNORECASE)
        clean_line = re.sub(r'\s+Revised.*$', '', clean_line, flags=re.IGNORECASE)
        
        if re.match(r'^\d+\s+[\w\s]+(?:Way|Street|St|Ave|Avenue|Road|Rd|Drive|Dr|Blvd|Crescent|Cres|Place|Pl|Court|Ct)', clean_line, re.IGNORECASE):
            extracted['client_address'] = clean_line
            print(f"✅ Client address: {extracted['client_address']}")
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                # Clean next line too
                next_line = re.sub(r'\s+(January|February|March|April|May|June|July|August|September|October|November|December).*$', '', next_line, flags=re.IGNORECASE)
                next_line = re.sub(r'\s+Revised.*$', '', next_line, flags=re.IGNORECASE)
                if re.match(r'^[A-Za-z\s]+,?\s*BC', next_line, re.IGNORECASE):
                    extracted['client_address'] += ', ' + next_line
                    print(f"✅ Updated address: {extracted['client_address']}")
            break
    
    # Extract phone
    phone_match = re.search(r'Tel[:\s]*(\d{3}[-.\s]?\d{3}[-.\s]?\d{4})', text, re.IGNORECASE)
    if phone_match:
        extracted['client_phone'] = phone_match.group(1).strip()
        print(f"✅ Phone: {extracted['client_phone']}")
    
    # Extract email
    email_match = re.search(r'Email[:\s]*([^\s\n]+@[^\s\n]+)', text, re.IGNORECASE)
    if email_match:
        email = email_match.group(1).strip()
        if 'allpaintingltd' not in email.lower():
            extracted['client_email'] = email
            print(f"✅ Email: {extracted['client_email']}")
    
    # Extract project name from "Project Reference:" section
    proj_match = re.search(r'Project Reference[:\s]*\n([^\n]+)', text, re.IGNORECASE)
    if proj_match:
        extracted['project_name'] = proj_match.group(1).strip()
        print(f"✅ Project name: {extracted['project_name']}")
    
    # Extract project address
    proj_ref_pos = text.find('Project Reference')
    if proj_ref_pos > 0:
        after_proj = text[proj_ref_pos:proj_ref_pos + 300]
        addr_match = re.search(r'\n(\d+\s+[\w\s]+(?:Way|Street|St|Ave|Avenue|Road|Rd|Drive|Dr|Blvd|Crescent|Cres|Place|Pl|Court|Ct)[^\n]*)', after_proj, re.IGNORECASE)
        if addr_match:
            extracted['project_address'] = addr_match.group(1).strip()
            print(f"✅ Project address: {extracted['project_address']}")
    
    # Extract scope items (Project Scope #1): description $ amount)
    scope_matches = re.findall(r'Project Scope\s*#?\d+\)?[:\s]*([^\$]+)\$\s*([\d,]+\.?\d*)', text, re.IGNORECASE)
    for desc, amount in scope_matches:
        desc = desc.strip().rstrip(':).').strip()
        amt = safe_float(amount)
        if desc and amt > 0:
            extracted['line_items'].append({
                'description': desc,
                'amount': amt
            })
            print(f"   • {desc}: ${amt:,.2f}")
    
    # Calculate subtotal
    extracted['subtotal'] = sum(item['amount'] for item in extracted['line_items'])
    print(f"\n💰 Subtotal: ${extracted['subtotal']:,.2f}")
    
    return extracted


def extract_new_format(text, page, extracted):
    """Extract data from new format estimates (from estimate generator app)"""
    
    # Extract estimate number
    est_match = re.search(r'Estimate[:\s]+(\d{2}C\s*[-–]\s*\d+)', text, re.IGNORECASE)
    if est_match:
        extracted['estimate_number'] = est_match.group(1).replace('–', '-').replace(' ', '')
        # Extract prefix (e.g., "25C") and base number
        prefix_match = re.search(r'(\d{2}C)-(\d+)', extracted['estimate_number'])
        if prefix_match:
            extracted['estimate_prefix'] = prefix_match.group(1)
            extracted['base_number'] = int(prefix_match.group(2))
        print(f"✅ Estimate number: {extracted['estimate_number']}")
        print(f"✅ Estimate prefix: {extracted.get('estimate_prefix', 'N/A')}")
    
    # Extract Company and Project
    company_project_match = re.search(r'Company:\s*(.+?)\s*Project:\s*(.+?)(?:\n|$)', text)
    if company_project_match:
        extracted['client_company'] = company_project_match.group(1).strip()
        extracted['project_name'] = company_project_match.group(2).strip()
        print(f"✅ Client company: {extracted['client_company']}")
        print(f"✅ Project name: {extracted['project_name']}")
    
    # Extract Contact
    contact_match = re.search(r'Contact:\s*([^\n]+?)(?:\s+Painting|\s+Address:|$)', text)
    if contact_match:
        extracted['client_contact'] = contact_match.group(1).strip()
        print(f"✅ Contact: {extracted['client_contact']}")
    
    # Extract addresses
    addr_match = re.search(r'Address:\s*(.+?)\s*Address:\s*(.+?)(?:\n|$)', text)
    if addr_match:
        extracted['client_address'] = addr_match.group(1).strip().rstrip(',')
        extracted['project_address'] = addr_match.group(2).strip()
        print(f"✅ Client address: {extracted['client_address']}")
        print(f"✅ Project address: {extracted['project_address']}")
    
    # Postal code
    postal_match = re.search(r'\n([A-Z]{2}\s+[A-Z]\d[A-Z]\s*\d[A-Z]\d)\n', text)
    if postal_match and extracted['client_address']:
        extracted['client_address'] += ', ' + postal_match.group(1).strip()
        print(f"✅ Updated client address: {extracted['client_address']}")
    
    # Phone - handle various formats including multiple numbers
    phone_match = re.search(r'Phone:\s*([\d\-\s/]+?)(?:\n|Email)', text)
    if phone_match:
        extracted['client_phone'] = phone_match.group(1).strip()
        print(f"✅ Client phone: {extracted['client_phone']}")
    
    # Email
    all_emails = re.findall(r'Email:\s*([^\s\n]+@[^\s\n]+)', text)
    for email in all_emails:
        if 'allpaintingltd' not in email.lower():
            extracted['client_email'] = email.strip()
            print(f"✅ Client email: {extracted['client_email']}")
            break
    
    # Extract from table
    tables = page.extract_tables()
    for table in tables:
        if table and len(table) > 1:
            header = table[0] if table[0] else []
            header_text = ' '.join(str(h) for h in header if h).lower()
            
            if 'scope' in header_text or 'description' in header_text or 'amount' in header_text:
                print(f"\n📋 Found Executive Summary table with {len(table)-1} rows")
                
                for row in table[1:]:
                    if row and len(row) >= 2:
                        row_text = ' '.join(str(cell) for cell in row if cell).lower()
                        if 'total' in row_text:
                            continue
                        
                        description = ''
                        amount = 0.0
                        
                        for cell in row:
                            if cell:
                                cell_str = str(cell).strip()
                                if '$' in cell_str or re.match(r'^[\d,]+\.?\d*$', cell_str.replace(',', '')):
                                    amount = safe_float(cell_str)
                                elif not cell_str.isdigit() and len(cell_str) > 2:
                                    if description:
                                        description += ' - ' + cell_str
                                    else:
                                        description = cell_str
                        
                        if description and amount > 0:
                            extracted['line_items'].append({
                                'description': description,
                                'amount': amount
                            })
                            print(f"   • {description}: ${amount:,.2f}")
    
    # Calculate subtotal
    extracted['subtotal'] = sum(item['amount'] for item in extracted['line_items'])
    print(f"\n💰 Subtotal: ${extracted['subtotal']:,.2f}")
    
    if extracted['subtotal'] == 0:
        total_match = re.search(r'TOTAL[^\$]*\$?([\d,]+\.?\d*)', text, re.IGNORECASE)
        if total_match:
            extracted['subtotal'] = safe_float(total_match.group(1))
            print(f"💰 Total from text: ${extracted['subtotal']:,.2f}")
    
    return extracted


def generate_invoice_html(invoice_data):
    """Generate the HTML for the invoice - MATCHING EXACT ESTIMATE FORMAT"""
    
    # Calculate totals
    subtotal = safe_float(invoice_data.get('subtotal', 0))
    gst_rate = 0.05
    gst = subtotal * gst_rate
    total = subtotal + gst
    
    # Format date
    invoice_date = invoice_data.get('date', datetime.now().strftime('%B %d, %Y'))
    
    # Logo HTML - matching estimate format exactly
    if HAS_LOGO:
        logo_html = f'''<div>
            <img src="file://{LOGO_URL}" style="height: 90px; width: auto; display: block; margin-bottom: 5pt;">
            <div style="font-size: 9pt; color: #555; line-height: 1.4;">
                Unit 1001, 115 Schoolhouse St, Coquitlam, BC, V3K 4X8<br>
                Phone: (604) 525-1403 | Email: info@allpaintingltd.com<br>
                www.allpaintingltd.com
            </div>
        </div>'''
    else:
        logo_html = '''<div>
            <div style="font-size: 22pt; font-weight: bold; color: #2B4C7E;">ALL PAINTING LTD.</div>
            <div style="font-size: 10pt; color: #555;">Your Full-Service Painting Company</div>
            <div style="font-size: 9pt; color: #555; margin-top: 5pt; line-height: 1.4;">
                Unit 1001, 115 Schoolhouse St, Coquitlam, BC, V3K 4X8<br>
                Phone: (604) 525-1403 | Email: info@allpaintingltd.com<br>
                www.allpaintingltd.com
            </div>
        </div>'''
    
    # Build scope items table rows (matching estimate executive summary format)
    scope_rows_html = ""
    for i, item in enumerate(invoice_data.get('line_items', []), 1):
        amount = safe_float(item.get('amount', 0))
        scope_rows_html += f'''
            <tr>
                <td style="padding: 10pt; border-bottom: 1px solid #ddd; font-size: 11pt;">{i}</td>
                <td style="padding: 10pt; border-bottom: 1px solid #ddd; font-size: 11pt;">{item.get('description', '')}</td>
                <td style="padding: 10pt; border-bottom: 1px solid #ddd; font-size: 11pt; text-align: right;">${amount:,.2f}</td>
            </tr>
        '''
    
    html = f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page {{
            size: letter;
            margin: 0.5in 0.5in 1.2in 0.5in;
            @bottom-center {{
                content: element(footer);
            }}
            @bottom-right {{
                content: "Page " counter(page);
                font-size: 10pt;
                color: #666;
            }}
        }}
        
        #footer {{
            position: running(footer);
            width: 100%;
        }}
        
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
            font-size: 11pt;
            line-height: 1.35;
            color: #333333;
        }}
        
        .no-break {{
            page-break-inside: avoid !important;
            break-inside: avoid !important;
        }}
    </style>
</head>
<body>
    <!-- Footer (running element) -->
    <div id="footer">
        <div style="border-top: 2px solid #2B4C7E; padding-top: 8pt; text-align: center; font-size: 10pt; color: #333;">
            ALL PAINTING LTD | Phone: (604) 525-1403 | www.allpaintingltd.com
        </div>
    </div>

    <!-- HEADER - Matching Estimate Format Exactly -->
    <div style="margin-bottom: 12pt; page-break-inside: avoid !important; border-bottom: 3px solid #2B4C7E; padding-bottom: 12pt;">
        <div style="display: flex; justify-content: space-between; align-items: flex-start;">
            {logo_html}
            <div style="text-align: right;">
                <div style="font-size: 22pt; font-weight: bold; color: #2B4C7E; margin-bottom: 4pt;">INVOICE</div>
                <div style="font-size: 11pt; color: #333;">Invoice: {invoice_data.get('invoice_number', '')}</div>
                <div style="font-size: 11pt; color: #333;">Date: {invoice_date}</div>
                <div style="font-size: 11pt; color: #333;">Reference: Est. {invoice_data.get('estimate_number', '')}</div>
            </div>
        </div>
    </div>
    
    <!-- CLIENT & PROJECT INFO - Two Boxes Side by Side (Matching Estimate) -->
    <div style="display: flex; gap: 12pt; margin-bottom: 12pt; page-break-inside: avoid !important;">
        <div style="flex: 1; border: 3px solid #2B4C7E; border-radius: 10px; padding: 12pt; background: white;">
            <div style="font-size: 12pt; font-weight: bold; color: #2B4C7E; margin-bottom: 8pt; border-bottom: 2px solid #2B4C7E; padding-bottom: 4pt;">CLIENT INFORMATION</div>
            <div style="font-size: 11pt; line-height: 1.4;">
                <div style="margin-bottom: 2pt;"><strong>Company:</strong> {invoice_data.get('client_company', '')}</div>
                <div style="margin-bottom: 2pt;"><strong>Contact:</strong> {invoice_data.get('client_contact', '')}</div>
                <div style="margin-bottom: 2pt;"><strong>Address:</strong> {invoice_data.get('client_address', '')}</div>
                <div style="margin-bottom: 2pt;"><strong>Phone:</strong> {invoice_data.get('client_phone', '')}</div>
                <div style="margin-bottom: 2pt;"><strong>Email:</strong> {invoice_data.get('client_email', '')}</div>
            </div>
        </div>
        <div style="flex: 1; border: 3px solid #2B4C7E; border-radius: 10px; padding: 12pt; background: white;">
            <div style="font-size: 12pt; font-weight: bold; color: #2B4C7E; margin-bottom: 8pt; border-bottom: 2px solid #2B4C7E; padding-bottom: 4pt;">PROJECT</div>
            <div style="font-size: 11pt; line-height: 1.4;">
                <div style="margin-bottom: 2pt;"><strong>Project:</strong> {invoice_data.get('project_name', '')}</div>
                <div style="margin-bottom: 2pt;"><strong>Address:</strong> {invoice_data.get('project_address', '')}</div>
            </div>
        </div>
    </div>
    
    <!-- EXECUTIVE SUMMARY - Matching Estimate Format -->
    <div class="no-break" style="margin-bottom: 12pt;">
        <div style="font-size: 14pt; font-weight: bold; color: #2B4C7E; margin-bottom: 8pt;">INVOICE SUMMARY</div>
        <table style="width: 100%; border-collapse: collapse; page-break-inside: avoid !important;">
            <thead>
                <tr style="background: #F5F5F5;">
                    <th style="padding: 10pt; text-align: left; font-size: 11pt; font-weight: bold; border-bottom: 2px solid #2B4C7E; width: 60pt;">Scope</th>
                    <th style="padding: 10pt; text-align: left; font-size: 11pt; font-weight: bold; border-bottom: 2px solid #2B4C7E;">Description</th>
                    <th style="padding: 10pt; text-align: right; font-size: 11pt; font-weight: bold; border-bottom: 2px solid #2B4C7E; width: 120pt;">Amount</th>
                </tr>
            </thead>
            <tbody>
                {scope_rows_html}
            </tbody>
        </table>
        
        <!-- Totals Section -->
        <table style="width: 100%; border-collapse: collapse; margin-top: 0;">
            <tr>
                <td style="padding: 10pt; font-size: 11pt; font-weight: bold;" colspan="2">Subtotal</td>
                <td style="padding: 10pt; font-size: 11pt; text-align: right; width: 120pt;">${subtotal:,.2f}</td>
            </tr>
            <tr>
                <td style="padding: 10pt; font-size: 11pt; font-weight: bold;" colspan="2">GST (5%)</td>
                <td style="padding: 10pt; font-size: 11pt; text-align: right; width: 120pt;">${gst:,.2f}</td>
            </tr>
            <tr style="background: #2B4C7E; color: white;">
                <td style="padding: 10pt; font-size: 12pt; font-weight: bold;" colspan="2">TOTAL (Including GST)</td>
                <td style="padding: 10pt; font-size: 12pt; font-weight: bold; text-align: right; width: 120pt;">${total:,.2f}</td>
            </tr>
        </table>
    </div>
    
    <!-- PAYMENT TERMS - Matching Estimate Format -->
    <div class="no-break" style="margin-bottom: 12pt;">
        <div style="display: flex; gap: 12pt; page-break-inside: avoid !important;">
            <div style="flex: 1; border: 3px solid #2B4C7E; border-radius: 10px; background: #F8F9FA; padding: 12pt;">
                <div style="font-size: 28pt; margin-bottom: 6pt; text-align: center;">💳</div>
                <div style="font-size: 12pt; font-weight: bold; color: #2B4C7E; margin-bottom: 6pt; text-align: center;">PAYMENT OPTIONS</div>
                <div style="font-size: 10pt; line-height: 1.4;">
                    <div style="margin-bottom: 8pt; text-align: center;"><strong style="color: #c00;">P.O. #:</strong> {invoice_data.get('po_number', '(Approved by email)')}</div>
                    <div style="margin-bottom: 5pt;"><strong>E-Transfer:</strong> info@allpaintingltd.com</div>
                    <div style="margin-top: 8pt; padding-top: 8pt; border-top: 1px solid #ddd;">
                        <strong>Bank Transfer:</strong><br>
                        <div style="margin-top: 4pt; font-size: 9pt;">
                            <table style="width: 100%; font-size: 9pt;">
                                <tr><td>Account Name:</td><td style="text-align: right;">ALL PAINTING LTD.</td></tr>
                                <tr><td>Transit No.:</td><td style="text-align: right;">94500</td></tr>
                                <tr><td>Institution No.:</td><td style="text-align: right;">004</td></tr>
                                <tr><td>Account No.:</td><td style="text-align: right;">5018919</td></tr>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
            <div style="flex: 1; border: 3px solid #2B4C7E; border-radius: 10px; background: #F8F9FA; padding: 12pt;">
                <div style="font-size: 28pt; margin-bottom: 6pt; text-align: center;">📋</div>
                <div style="font-size: 12pt; font-weight: bold; color: #2B4C7E; margin-bottom: 6pt; text-align: center;">PAYMENT TERMS</div>
                <div style="font-size: 10pt; line-height: 1.4;">
                    <ul style="margin: 0; padding-left: 20pt; list-style-type: disc;">
                        <li style="margin-bottom: 3pt;">Payment due upon receipt</li>
                        <li style="margin-bottom: 3pt;">Net 30 days</li>
                    </ul>
                </div>
            </div>
        </div>
    </div>
    
    <!-- Thank You -->
    <div style="margin-top: 20pt; text-align: center;">
        <div style="font-size: 11pt; margin-bottom: 10pt;">Thank you for your business.</div>
        <div style="font-weight: bold; font-size: 12pt; color: #2B4C7E;">ALL PAINTING LTD.</div>
    </div>
</body>
</html>'''
    
    return html


def generate_pdf_from_html(html_content, output_path):
    """Convert HTML to PDF using WeasyPrint"""
    extra_css = CSS(string='''
        @page {
            size: letter;
            margin: 0.5in;
        }
    ''')
    
    HTML(string=html_content).write_pdf(output_path, stylesheets=[extra_css])
    return output_path


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory('static', filename)


@app.route('/api/counter', methods=['GET'])
def get_counter():
    counter = load_invoice_counter()
    return jsonify(counter)


@app.route('/api/counter/reset', methods=['POST'])
def reset_counter():
    """Reset the invoice counter - can reset specific prefix or all"""
    try:
        data = request.json or {}
        prefix = data.get('prefix')
        
        counter = load_invoice_counter()
        
        if prefix:
            # Reset specific prefix sequence
            if prefix in counter.get('sequences', {}):
                del counter['sequences'][prefix]
                save_invoice_counter(counter)
                return jsonify({
                    'success': True,
                    'message': f'Reset sequence for {prefix}. Next invoice will be {prefix}-001'
                })
            else:
                return jsonify({
                    'success': True,
                    'message': f'No sequence found for {prefix}'
                })
        else:
            # Reset all sequences
            counter['sequences'] = {}
            save_invoice_counter(counter)
            return jsonify({
                'success': True,
                'message': 'All invoice sequences reset'
            })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/counter/set', methods=['POST'])
def set_counter():
    """Set a specific sequence number for a prefix"""
    try:
        data = request.json
        prefix = data.get('prefix')
        sequence = data.get('sequence', 0)
        
        if not prefix:
            return jsonify({'error': 'prefix is required'}), 400
        
        counter = load_invoice_counter()
        sequences = counter.get('sequences', {})
        sequences[prefix] = int(sequence)
        counter['sequences'] = sequences
        save_invoice_counter(counter)
        
        next_num = int(sequence) + 1
        return jsonify({
            'success': True,
            'message': f'Set {prefix} sequence to {sequence}. Next invoice will be {prefix}-{next_num:03d}'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/upload', methods=['POST'])
def upload_estimate():
    """Upload estimate PDF and extract data"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Only PDF files are supported'}), 400
    
    try:
        # Save uploaded file
        filename = f"estimate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Extract data from PDF
        extracted_data = extract_data_from_estimate_pdf(filepath)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'extracted_data': extracted_data,
            'message': 'Estimate data extracted successfully'
        })
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/generate', methods=['POST'])
def generate_invoice():
    """Generate an invoice PDF from the extracted/edited data"""
    try:
        invoice_data = request.json
        
        # Generate invoice number based on estimate prefix
        estimate_prefix = invoice_data.get('estimate_prefix', '')
        invoice_data['invoice_number'] = get_next_invoice_number(estimate_prefix)
        
        # Keep estimate number for reference
        base_number = invoice_data.get('base_number', 0)
        if not invoice_data.get('estimate_number'):
            prefix = estimate_prefix or get_current_year_prefix()
            invoice_data['estimate_number'] = f"{prefix}-{int(base_number):03d}"
        
        print(f"\n{'='*60}")
        print(f"GENERATING INVOICE: {invoice_data['invoice_number']}")
        print(f"Reference Estimate: {invoice_data['estimate_number']}")
        print(f"{'='*60}")
        
        # Generate HTML
        html_content = generate_invoice_html(invoice_data)
        
        # Generate PDF filename
        safe_client = invoice_data.get('client_company', 'Client').replace('/', '-').replace('\\', '-')[:30]
        safe_project = invoice_data.get('project_name', 'Project').replace('/', '-').replace('\\', '-')[:30]
        invoice_num_safe = invoice_data['invoice_number'].replace(' ', '_').replace('(', '').replace(')', '')
        
        output_filename = f"Invoice_{invoice_num_safe}_{safe_client}.pdf"
        for char in ['#', '&', '%', '?', ':', '*', '"', '<', '>', '|']:
            output_filename = output_filename.replace(char, '-')
        
        output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        
        # Generate PDF
        generate_pdf_from_html(html_content, output_path)
        
        print(f"✅ PDF generated: {output_filename}")
        print(f"{'='*60}\n")
        
        return jsonify({
            'success': True,
            'pdf_filename': output_filename,
            'invoice_number': invoice_data['invoice_number'],
            'message': 'Invoice generated successfully'
        })
    
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(filepath, as_attachment=True, download_name=filename)

"""
Additional API endpoints for Invoice Generator - Automation Integration
Add this code to the end of app.py (before if __name__ == '__main__':)
"""

# =============================================================================
# CONFIGURATION FOR AUTOMATION
# =============================================================================

INVOICE_BASE_FOLDER = r"\\Aps\apl\Invoices"
EXCEL_PATH = r"N:\Estimates\Estimate_InvoiceNumbers\2026 Master Estimates_Invoice List.xlsx"
EXCEL_SHEET_NAME = "Invoices 2026"

# Outlook Draft Settings
DRAFTS_FOLDER_NAME = "Drafts - Invoices"


def get_invoice_output_folder(estimate_prefix):
    """
    Get the correct invoice output folder based on the estimate prefix year.
    
    Args:
        estimate_prefix: Prefix like '25C', '26R', etc.
    
    Returns:
        Full path to the invoice folder for that year
    """
    # Extract year from prefix (first 2 digits)
    if estimate_prefix and len(estimate_prefix) >= 2:
        year_prefix = estimate_prefix[:2]
        year_full = f"20{year_prefix}"
    else:
        # Default to current year
        year_full = str(datetime.now().year)
        year_prefix = year_full[-2:]
    
    folder_name = f"{year_full} All Painting Invoices"
    return os.path.join(INVOICE_BASE_FOLDER, folder_name)


def get_next_invoice_number_from_excel():
    """
    Get the next invoice number from the Excel spreadsheet.
    Finds the first row where Customer (column C) is empty and returns the Invoice # from column A.
    
    Returns:
        Tuple of (invoice_number, row_number) or (None, None) if not found
    """
    import openpyxl
    
    if not os.path.exists(EXCEL_PATH):
        print(f"⚠️ Excel file not found: {EXCEL_PATH}")
        return None, None
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[EXCEL_SHEET_NAME]
        
        # Find the first row where Customer (column C) is empty
        row = 2  # Start after header
        while row < 1000:  # Safety limit
            customer = ws.cell(row=row, column=3).value
            invoice_num = ws.cell(row=row, column=1).value
            
            if (customer is None or customer == '') and invoice_num:
                print(f"📊 Found available invoice number: {invoice_num} at row {row}")
                wb.close()
                return str(invoice_num), row
            
            row += 1
        
        wb.close()
        print(f"⚠️ No available invoice numbers found in Excel")
        return None, None
        
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        return None, None


def get_next_invoice_number_from_folder(estimate_prefix=None):
    """
    Get the next invoice number - NOW READS FROM EXCEL FIRST.
    Falls back to folder scanning if Excel doesn't work.
    
    Args:
        estimate_prefix: Prefix like '25C', '26R', etc. (not used anymore, kept for compatibility)
    
    Returns:
        Next invoice number string (e.g., '26-005')
    """
    import os
    import re
    
    # First, try to get from Excel
    excel_invoice_num, excel_row = get_next_invoice_number_from_excel()
    if excel_invoice_num:
        return excel_invoice_num
    
    # Fallback: scan folder (old method)
    print(f"⚠️ Falling back to folder scan method...")
    
    if not estimate_prefix:
        estimate_prefix = get_current_year_prefix()
    
    # Get the correct folder for this year
    invoice_folder = get_invoice_output_folder(estimate_prefix)
    
    # Scan the invoice folder for existing invoices with this prefix
    max_number = 0
    
    if os.path.exists(invoice_folder):
        for filename in os.listdir(invoice_folder):
            if filename.lower().endswith('.pdf'):
                # Look for pattern like Invoice_25C-003 or 25C-003
                match = re.search(rf'{re.escape(estimate_prefix)}-(\d+)', filename, re.IGNORECASE)
                if match:
                    num = int(match.group(1))
                    if num > max_number:
                        max_number = num
    
    # Also check the counter file to be safe
    counter = load_invoice_counter()
    sequences = counter.get('sequences', {})
    counter_num = sequences.get(estimate_prefix, 0)
    
    # Use the higher of the two
    max_number = max(max_number, counter_num)
    
    next_number = max_number + 1
    
    # Update counter file
    sequences[estimate_prefix] = next_number
    counter['sequences'] = sequences
    save_invoice_counter(counter)
    
    return f"{estimate_prefix}-{next_number:03d}"


def update_excel_with_invoice(invoice_data):
    """
    Update the Excel spreadsheet with the new invoice information.
    Does NOT overwrite the Invoice # column (A) since it's pre-filled.
    
    Args:
        invoice_data: Dictionary with invoice details
    """
    import openpyxl
    from datetime import datetime
    
    if not os.path.exists(EXCEL_PATH):
        print(f"⚠️ Excel file not found: {EXCEL_PATH}")
        return False
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[EXCEL_SHEET_NAME]
        
        # Find the row that matches our invoice number, or the first empty row
        invoice_number = invoice_data.get('invoice_number', '')
        target_row = None
        
        # First, try to find the row with our invoice number
        for row in range(2, 1000):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value) == str(invoice_number):
                target_row = row
                break
        
        # If not found, find first empty row (where Customer is empty)
        if target_row is None:
            target_row = 2
            while ws.cell(row=target_row, column=3).value is not None and ws.cell(row=target_row, column=3).value != '':
                target_row += 1
        
        print(f"📊 Updating invoice in Excel row {target_row}")
        
        # Column mapping:
        # A (1): Invoice # - DO NOT OVERWRITE (already pre-filled)
        # C (3): Customer
        # E (5): Job # (estimate number)
        # F (6): Subtotal
        # G (7): GST (formula)
        # H (8): Total (formula)
        # I (9): Job Type
        # J (10): Job Details
        # K (11): Project Address
        # L (12): Date
        # M (13): Invoice Notes
        
        # NOTE: We do NOT update column A (Invoice #) - it's already there
        ws.cell(row=target_row, column=3, value=invoice_data.get('client_company', '') or invoice_data.get('client_contact', ''))
        ws.cell(row=target_row, column=5, value=invoice_data.get('estimate_number', ''))
        ws.cell(row=target_row, column=6, value=invoice_data.get('subtotal', 0))
        # Don't overwrite formulas in columns G and H if they already exist
        if not ws.cell(row=target_row, column=7).value:
            ws.cell(row=target_row, column=7, value=f"=+F{target_row}*0.05")  # GST formula
        if not ws.cell(row=target_row, column=8).value:
            ws.cell(row=target_row, column=8, value=f"=+F{target_row}+G{target_row}")  # Total formula
        
        # Determine job type based on prefix
        estimate_prefix = invoice_data.get('estimate_prefix', '')
        if 'R' in estimate_prefix.upper():
            job_type = 'Residential'
        else:
            job_type = 'Commercial'
        ws.cell(row=target_row, column=9, value=job_type)
        
        ws.cell(row=target_row, column=10, value=invoice_data.get('project_name', ''))
        ws.cell(row=target_row, column=11, value=invoice_data.get('project_address', ''))
        ws.cell(row=target_row, column=12, value=datetime.now().strftime('%Y-%m-%d'))
        # ws.cell(row=target_row, column=13, value='')  # Don't clear Invoice Notes
        
        wb.save(EXCEL_PATH)
        print(f"✅ Excel updated successfully")
        return True
        
    except Exception as e:
        print(f"❌ Error updating Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def generate_invoice_filename(estimate_pdf_filename, invoice_number, estimate_number):
    """
    Generate the invoice filename based on estimate filename
    
    Example:
        Input estimate: "Estimate - 25C-325 - VC -One Guilford -14957,14969 100 Ave Surrey.pdf"
        Invoice number: "25C-004"
        Output: "Invoice - 26-005 - 25C-325 - VC -One Guilford -14957,14969 100 Ave Surrey.pdf"
    
    Args:
        estimate_pdf_filename: Original estimate PDF filename
        invoice_number: Invoice number from Excel (e.g., '26-005')
        estimate_number: Original estimate number (e.g., '25C-325')
    
    Returns:
        Invoice filename
    """
    import re
    
    # Remove "Estimate - " prefix and the estimate number
    # Pattern: "Estimate - 25C-325" or "Estimate - 25C325"
    cleaned = re.sub(r'^Estimate\s*-\s*\d{2}[A-Za-z]-?\d+(-[A-Za-z])?\s*', '', estimate_pdf_filename)
    
    # Clean up any leading dashes or spaces
    cleaned = cleaned.lstrip('- ')
    
    # Build new filename: Invoice - [invoice_number] - [estimate_number] - [rest of name]
    # Example: Invoice - 26-005 - 25C-319 - Keller Const - Int ptg - Shell North Surrey.pdf
    if cleaned:
        invoice_filename = f"Invoice - {invoice_number} - {estimate_number} - {cleaned}"
    else:
        invoice_filename = f"Invoice - {invoice_number} - {estimate_number}.pdf"
    
    # Ensure it ends with .pdf
    if not invoice_filename.lower().endswith('.pdf'):
        invoice_filename += '.pdf'
    
    return invoice_filename


def create_invoice_email_draft(invoice_data, invoice_path):
    """
    Create a professional email draft for sending the invoice
    
    Args:
        invoice_data: Dictionary with invoice and client details
        invoice_path: Path to the invoice PDF
    
    Returns:
        Dictionary with draft creation result
    """
    import requests
    import base64
    
    # Get client email
    client_email = invoice_data.get('client_email', '')
    if not client_email or '@' not in client_email:
        print(f"⚠️ No client email found, skipping draft creation")
        return {"success": False, "error": "No client email found"}
    
    # Handle multiple emails
    if '/' in client_email:
        client_email = client_email.split('/')[0].strip()
    
    # Get contact name for greeting
    contact_name = invoice_data.get('client_contact', '')
    if '/' in contact_name:
        contact_name = contact_name.split('/')[0].strip()
    first_name = contact_name.split()[0] if contact_name else ''
    
    # Prepare greeting
    greeting = f"Hi {first_name}," if first_name else "Hello,"
    
    # Get invoice details
    invoice_number = invoice_data.get('invoice_number', '')
    estimate_number = invoice_data.get('estimate_number', '')
    project_name = invoice_data.get('project_name', '')
    project_address = invoice_data.get('project_address', '')
    subtotal = invoice_data.get('subtotal', 0)
    gst = subtotal * 0.05
    total = subtotal + gst
    
    # Build project description
    if project_name and project_address:
        project_desc = f"the {project_name} project at {project_address}"
    elif project_name:
        project_desc = f"the {project_name} project"
    elif project_address:
        project_desc = f"the project at {project_address}"
    else:
        project_desc = "your project"
    
    # Email subject - same as invoice filename without .pdf
    subject = f"Invoice_{invoice_number} - {estimate_number}"
    if project_name:
        subject += f" - {project_name[:50]}"
    
    # Professional email body
    body_html = f"""
<html>
<body style="font-family: Arial, sans-serif; font-size: 14px; color: #333;">
<p>{greeting}</p>

<p>Please find attached Invoice <strong>{invoice_number}</strong> for {project_desc}, referencing Estimate {estimate_number}.</p>

<p><strong>Invoice Summary:</strong><br>
Subtotal: ${subtotal:,.2f}<br>
GST (5%): ${gst:,.2f}<br>
<strong>Total Due: ${total:,.2f}</strong></p>

<p>Please let us know if you have any questions regarding this invoice.</p>

<p>We appreciate your business and thank you for choosing All Painting Ltd.</p>

<p>Best regards,</p>

<p><strong>All Painting Ltd.</strong><br>
Phone: (604) 525-1403<br>
Email: info@allpaintingltd.com<br>
www.allpaintingltd.com</p>
</body>
</html>
"""
    
    # Call Outlook API to create draft
    try:
        # Import the outlook client from estimate-automation
        import sys
        sys.path.insert(0, r'C:\Scripts\estimate-automation')
        from outlook_client import OutlookClient
        
        outlook = OutlookClient()
        
        # Get or create the invoices drafts folder
        folder_id = outlook.get_or_create_folder(DRAFTS_FOLDER_NAME)
        
        # Create draft with attachment
        draft = outlook.create_draft_with_attachment(
            to_email=client_email,
            subject=subject,
            body_html=body_html,
            attachment_path=invoice_path,
            folder_id=folder_id
        )
        
        print(f"✅ Invoice email draft created")
        print(f"   📬 To: {client_email}")
        print(f"   📎 Subject: {subject}")
        
        return {
            "success": True,
            "draft_id": draft.get("id"),
            "to_email": client_email,
            "subject": subject
        }
        
    except Exception as e:
        print(f"⚠️ Could not create email draft: {e}")
        return {"success": False, "error": str(e)}


@app.route('/api/generate-from-estimate', methods=['POST'])
def api_generate_from_estimate():
    """
    API endpoint for automation - generates invoice from estimate PDF path
    
    Request JSON:
    {
        "estimate_pdf_path": "N:\\Estimates\\2026 All Painting Estimates\\Estimate - 25C-325 - Project.pdf"
    }
    
    Response JSON:
    {
        "success": true,
        "invoice_number": "25C-004",
        "invoice_path": "\\\\Aps\\apl\\Invoices\\2026 All Painting Invoices\\Invoice_25C-004 - 25C-325 - Project.pdf",
        "excel_updated": true,
        "draft_created": true,
        "draft_to_email": "client@email.com"
    }
    """
    try:
        data = request.get_json()
        
        if not data or 'estimate_pdf_path' not in data:
            return jsonify({
                'success': False,
                'error': 'Missing required field: estimate_pdf_path'
            }), 400
        
        estimate_pdf_path = data['estimate_pdf_path']
        
        # Validate file exists
        if not os.path.exists(estimate_pdf_path):
            return jsonify({
                'success': False,
                'error': f'Estimate PDF not found: {estimate_pdf_path}'
            }), 404
        
        print("\n" + "="*60)
        print("API AUTOMATION - GENERATING INVOICE FROM ESTIMATE")
        print("="*60)
        print(f"📁 Estimate PDF: {estimate_pdf_path}")
        
        # Step 1: Extract data from estimate PDF
        print(f"\n📄 Step 1: Extracting data from estimate PDF...")
        extracted_data = extract_data_from_estimate_pdf(estimate_pdf_path)
        
        # If estimate number not found in PDF, try to extract from filename
        estimate_filename = os.path.basename(estimate_pdf_path)
        if not extracted_data.get('estimate_number') or not extracted_data.get('estimate_prefix'):
            print(f"   ⚠️ Estimate number not found in PDF, extracting from filename...")
            filename_match = re.search(r'(\d{2}[A-Za-z])-?(\d+)', estimate_filename)
            if filename_match:
                extracted_data['estimate_prefix'] = filename_match.group(1).upper()
                extracted_data['estimate_number'] = f"{filename_match.group(1).upper()}-{filename_match.group(2)}"
                print(f"   ✅ Extracted from filename: {extracted_data['estimate_number']}")
        
        if not extracted_data.get('estimate_number') and not extracted_data.get('estimate_prefix'):
            return jsonify({
                'success': False,
                'error': 'Could not extract estimate number from PDF or filename'
            }), 500
        
        # Step 2: Get next invoice number
        estimate_prefix = extracted_data.get('estimate_prefix', get_current_year_prefix())
        estimate_number = extracted_data.get('estimate_number', '')
        print(f"\n🔢 Step 2: Getting next invoice number for prefix {estimate_prefix}...")
        invoice_number = get_next_invoice_number_from_folder(estimate_prefix)
        print(f"   Invoice number: {invoice_number}")
        
        # Step 3: Prepare invoice data
        invoice_data = {
            'invoice_number': invoice_number,
            'estimate_number': estimate_number,
            'estimate_prefix': estimate_prefix,
            'date': datetime.now().strftime('%B %d, %Y'),
            'client_company': extracted_data.get('client_company', ''),
            'client_contact': extracted_data.get('client_contact', ''),
            'client_address': extracted_data.get('client_address', ''),
            'client_phone': extracted_data.get('client_phone', ''),
            'client_email': extracted_data.get('client_email', ''),
            'project_name': extracted_data.get('project_name', ''),
            'project_address': extracted_data.get('project_address', ''),
            'line_items': extracted_data.get('line_items', []),
            'subtotal': extracted_data.get('subtotal', 0)
        }
        
        # Step 4: Generate invoice HTML and PDF
        print(f"\n📝 Step 3: Generating invoice PDF...")
        html_content = generate_invoice_html(invoice_data)
        
        # Generate filename with new format: Invoice_[invoice#] - [estimate#] - [rest]
        estimate_filename = os.path.basename(estimate_pdf_path)
        invoice_filename = generate_invoice_filename(estimate_filename, invoice_number, estimate_number)
        
        # Get the correct output folder based on year prefix
        invoice_output_folder = get_invoice_output_folder(estimate_prefix)
        
        # Ensure output folder exists
        if not os.path.exists(invoice_output_folder):
            os.makedirs(invoice_output_folder, exist_ok=True)
        
        invoice_path = os.path.join(invoice_output_folder, invoice_filename)
        
        print(f"   Saving to: {invoice_path}")
        generate_pdf_from_html(html_content, invoice_path)
        print(f"   ✅ Invoice PDF generated")
        
        # Step 5: Update Excel spreadsheet
        print(f"\n📊 Step 4: Updating Excel spreadsheet...")
        excel_updated = update_excel_with_invoice(invoice_data)
        
        # Step 6: Create email draft
        print(f"\n📧 Step 5: Creating email draft...")
        draft_result = create_invoice_email_draft(invoice_data, invoice_path)
        
        print("\n" + "="*60)
        print("API AUTOMATION - INVOICE GENERATION COMPLETE!")
        print("="*60 + "\n")
        
        response = {
            'success': True,
            'invoice_number': invoice_number,
            'invoice_path': invoice_path,
            'invoice_filename': invoice_filename,
            'estimate_number': estimate_number,
            'client': invoice_data.get('client_company') or invoice_data.get('client_contact', ''),
            'subtotal': invoice_data.get('subtotal', 0),
            'excel_updated': excel_updated,
            'message': 'Invoice generated successfully'
        }
        
        # Add draft info if created
        if draft_result.get('success'):
            response['draft_created'] = True
            response['draft_to_email'] = draft_result.get('to_email')
            response['draft_subject'] = draft_result.get('subject')
        else:
            response['draft_created'] = False
            response['draft_error'] = draft_result.get('error')
        
        return jsonify(response)
    
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"\n❌ API ERROR:\n{error_trace}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/health', methods=['GET'])
def api_health():
    """Health check endpoint"""
    current_prefix = get_current_year_prefix()
    current_folder = get_invoice_output_folder(current_prefix)
    return jsonify({
        'status': 'healthy',
        'service': 'invoice-generator',
        'logo_available': HAS_LOGO,
        'invoice_base_folder': INVOICE_BASE_FOLDER,
        'current_year_folder': current_folder,
        'invoice_folder_exists': os.path.exists(current_folder),
        'excel_exists': os.path.exists(EXCEL_PATH)
    })


# =============================================================================
# END OF ADDITIONAL CODE
# =============================================================================

if __name__ == '__main__':
    print("\n" + "="*60)
    print("ALL PAINTING LTD - INVOICE GENERATOR")
    print("Upload Estimate PDF → Generate Invoice")
    print("="*60)
    if HAS_LOGO:
        print(f"✅ Logo found: {LOGO_PATH}")
    else:
        print(f"⚠️  Logo not found at: {LOGO_PATH}")
    
    counter = load_invoice_counter()
    print(f"📊 Year prefix: {counter['year_prefix']}")
    print(f"📊 Sequences tracked: {len(counter.get('sequences', {}))}")
    print(f"📁 Invoice base folder: {INVOICE_BASE_FOLDER}")
    print(f"📊 Excel file: {EXCEL_PATH}")
    
    print("Server starting on http://localhost:5001")
    print("="*60 + "\n")
    app.run(host='0.0.0.0', port=5001, debug=True)
